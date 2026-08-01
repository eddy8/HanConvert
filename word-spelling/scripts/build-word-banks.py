#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORDS_DIR = ROOT / "words"
OUTPUT_PATH = WORDS_DIR / "word-banks.json"

WORD_HEADERS = {"单词", "英文单词原形", "word", "english"}
PHONETIC_HEADERS = {"音标", "国际音标", "ipa", "美音", "英音"}
MEANING_HEADERS = {"释义", "中文释义", "含义", "meaning"}
EXAMPLE_HEADERS = {"例句", "例句 (可选)", "例句或用法说明", "example"}


def clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def header_key(value) -> str:
    return clean_text(value).lower()


def normalize_phonetic(value: str) -> str:
    phonetic = clean_text(value)
    if not phonetic:
        return ""
    if phonetic.startswith("[") and phonetic.endswith("]"):
        return f"/{phonetic[1:-1]}/"
    return phonetic


def pick_column(header_map: dict[str, int], names: set[str]) -> int | None:
    for name in names:
        key = name.lower()
        if key in header_map:
            return header_map[key]
    return None


def parse_workbook(path: Path, index: int) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    header = next(rows, ())
    header_map = {header_key(value): idx for idx, value in enumerate(header) if clean_text(value)}

    word_col = pick_column(header_map, WORD_HEADERS)
    meaning_col = pick_column(header_map, MEANING_HEADERS)
    example_col = pick_column(header_map, EXAMPLE_HEADERS)
    phonetic_col = pick_column(header_map, {"音标", "国际音标", "ipa", "美音"}) or pick_column(header_map, {"英音"})

    if word_col is None:
        raise ValueError(f"{path.name} 缺少单词列")

    words: list[dict[str, str]] = []
    seen: set[str] = set()

    for row in rows:
        raw_word = clean_text(row[word_col] if word_col < len(row) else "")
        if not raw_word or not re.fullmatch(r"[A-Za-z]+", raw_word):
            continue

        key = raw_word.lower()
        if key in seen:
            continue
        seen.add(key)

        item = {
            "word": raw_word,
            "phonetic": normalize_phonetic(row[phonetic_col] if phonetic_col is not None and phonetic_col < len(row) else ""),
            "meaning": clean_text(row[meaning_col] if meaning_col is not None and meaning_col < len(row) else ""),
            "example": clean_text(row[example_col] if example_col is not None and example_col < len(row) else ""),
        }
        words.append(item)

    return {
        "id": f"bank-{index:03d}",
        "label": path.stem,
        "source": path.name,
        "count": len(words),
        "words": words,
    }


def main() -> None:
    banks = []
    for index, path in enumerate(sorted(WORDS_DIR.glob("*.xlsx")), start=1):
        if path.name.startswith(".~") or path.name.startswith("~$"):
            continue
        bank = parse_workbook(path, len(banks) + 1)
        if bank["words"]:
            banks.append(bank)

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "sourceDirectory": "words",
        "banks": banks,
    }
    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH} with {len(banks)} banks.")


if __name__ == "__main__":
    main()
